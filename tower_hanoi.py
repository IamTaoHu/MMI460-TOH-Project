import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import time
import csv
import colorsys

class HanoiGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Tower of Hanoi - GUI (Drag & Drop)")
        self.root.geometry("1200x520")
        self.root.minsize(1200, 520)

        # ---- Theme ----
        self.bg = "#0f0f13"
        self.fg = "#d7e3ff"
        self.accent = "#8ab4ff"
        self.peg_color = "#ff5f5f"
        self.base_color = "#ff6d6d"
        self.move_fg = "#8be9fd"
        self.btn_bg = "#2a2f3a"
        self.btn_hover = "#394153"
        self.timer_fg = "#ffd86b"

        self.root.configure(bg=self.bg)

        # ---- Game state ----
        self.num_discs = 3
        self.max_discs = 8
        self.min_discs = 3
        self.moves = 0
        self.pegs = [[], [], []]  # bottom -> top
        self.disc_sizes = {}
        self.palette = []  # generated distinct colors per game
        # Drag & drop
        self.dragging_disc = None
        self.drag_from_peg = None
        self.drag_offset_x = 0
        self.drag_offset_y = 0
        self.interactions_enabled = True
        self.solving = False
        self.rule_break_attempts = 0

        # --- Move logging (time since first click) ---
        self.first_click_baseline = None
        self.current_move_start_rel = None
        self.move_logs = []  # (start_rel, end_rel, from_peg(1-3), to_peg(1-3))
        self.completed_elapsed_ms = None
        self.log_window = None
        self.log_text = None

        # --- Countdown timer ---
        self.timer_minutes = 0          # 0..99; 0 = disabled
        self.timer_seconds_left = None  # int seconds; None = disabled/not started
        self.timer_job = None
        self.game_over = False

        # --- Player records / table ---
        self.records = []      # dict: Name, Num of Disc, Move, Breaking rules, Timer, Remaining time
        self.table_window = None
        self.table_tree = None

        self._build_ui()
        self._new_game()

    # ---------- Color palette ----------

    def _gen_palette(self, n):
        """Generate n distinct PASTEL colors for dark background."""
        colors = []
        w = 0.18  # mix-with-white factor for extra pastel softness
        for i in range(max(1, n)):
            h = i / max(1, n)        # spread hues evenly
            s = 0.28                 # lower saturation = pastel
            v = 1.00                 # bright to stand out on dark bg
            import colorsys
            r, g, b = colorsys.hsv_to_rgb(h, s, v)
            # Mix slightly with white to soften further
            r = r*(1-w) + 1*w
            g = g*(1-w) + 1*w
            b = b*(1-w) + 1*w
            colors.append(f"#{int(r*255):02x}{int(g*255):02x}{int(b*255):02x}")
        return colors

    # ---------- UI ----------
    def _build_ui(self):
        top = tk.Frame(self.root, bg=self.bg)
        top.pack(fill="x", padx=12, pady=8)

        # Disks control
        tk.Label(top, text="Disks:", bg=self.bg, fg=self.fg, font=("Segoe UI", 12, "bold")).pack(side="left")
        self.disks_label_val = tk.Label(top, text=str(self.num_discs), bg=self.bg, fg=self.fg, font=("Segoe UI", 12, "bold"))
        minus = self._mk_btn(top, "▼", lambda: self._change_disks(-1))
        plus  = self._mk_btn(top, "▲", lambda: self._change_disks(+1))
        self.disks_label_val.pack(side="left", padx=(6, 6))
        minus.pack(side="left", padx=(0,4))
        plus.pack(side="left", padx=(0,16))

        # Moves label
        tk.Label(top, text="Moves:", bg=self.bg, fg=self.fg, font=("Segoe UI", 12, "bold")).pack(side="left")
        self.moves_label = tk.Label(top, text="0", bg=self.bg, fg=self.move_fg, font=("Consolas", 14, "bold"))
        self.moves_label.pack(side="left", padx=(6, 16))

        tk.Label(top, text="Breaking rules:", bg=self.bg, fg=self.fg, font=("Segoe UI", 12, "bold")).pack(side="left")
        self.break_label = tk.Label(top, text="0", bg=self.bg, fg=self.timer_fg, font=("Consolas", 14, "bold"))
        self.break_label.pack(side="left", padx=(6, 16))

        # Buttons
        self.btn_restart = self._mk_btn(top, "Restart", self._new_game)
        self.btn_log = self._mk_btn(top, "Log", self._open_log_window)
        self.btn_solve = self._mk_btn(top, "Solve!", self._solve_animate)
        self.btn_restart.pack(side="left", padx=4)
        self.btn_log.pack(side="left", padx=4)
        self.btn_solve.pack(side="left", padx=4)

        # Timer controls (right side)
        spacer = tk.Frame(top, bg=self.bg)
        spacer.pack(side="left", expand=True, fill="x")
        tk.Label(top, text="Timer (min):", bg=self.bg, fg=self.fg, font=("Segoe UI", 12, "bold")).pack(side="left", padx=(0,8))
        self.btn_timer_minus = self._mk_btn(top, "▼", lambda: self._change_timer_minutes(-1))
        self.timer_min_label = tk.Label(top, text=str(self.timer_minutes), bg=self.bg, fg=self.fg, font=("Segoe UI", 12, "bold"))
        self.btn_timer_plus  = self._mk_btn(top, "▲", lambda: self._change_timer_minutes(+1))
        self.btn_timer_minus.pack(side="left", padx=(0,4))
        self.timer_min_label.pack(side="left", padx=(0,4))
        self.btn_timer_plus.pack(side="left", padx=(0,16))

        self.timer_label = tk.Label(top, text="00:00", bg=self.bg, fg=self.timer_fg, font=("Consolas", 16, "bold"))
        self.timer_label.pack(side="left", padx=(0,16))

        # Name Entry (ข้างๆ Table ตามสเปค)
        tk.Label(top, text="Name:", bg=self.bg, fg=self.fg, font=("Segoe UI", 12, "bold")).pack(side="left", padx=(0,6))
        self.name_var = tk.StringVar()
        self.name_entry = tk.Entry(top, textvariable=self.name_var, width=18, bg="#1a1f29", fg=self.fg,
                                   insertbackground=self.fg, relief="flat", font=("Segoe UI", 11))
        self.name_entry.pack(side="left")

        # Canvas
        self.canvas = tk.Canvas(self.root, bg=self.bg, highlightthickness=0)
        self.canvas.pack(fill="both", expand=True, padx=12, pady=4)

        # Bottom: Save / Table + Min moves
        bottom = tk.Frame(self.root, bg=self.bg)
        bottom.pack(fill="x", padx=12, pady=(0,10))
        self.btn_save = self._mk_btn(bottom, "Save", self._save_record)
        self.btn_table = self._mk_btn(bottom, "Table", self._open_table_window)
        self.btn_save.pack(side="left", padx=(0,6))
        self.btn_table.pack(side="left", padx=(0,6))

        self.min_moves_label = tk.Label(bottom, text="", bg=self.bg, fg=self.fg, font=("Segoe UI", 11))
        self.min_moves_label.pack(side="right")

        # Mouse bindings
        self.canvas.bind("<Button-1>", self._on_mouse_down)
        self.canvas.bind("<B1-Motion>", self._on_mouse_move)
        self.canvas.bind("<ButtonRelease-1>", self._on_mouse_up)

        # Resize handling
        self.canvas.bind("<Configure>", lambda e: self._redraw())

    def _mk_btn(self, parent, text, cmd):
        return tk.Button(parent, text=text, command=cmd, bg=self.btn_bg, fg=self.fg,
                         activebackground=self.btn_hover, activeforeground=self.fg,
                         bd=0, relief="flat", font=("Segoe UI", 11, "bold"),
                         padx=10, pady=4, cursor="hand2")

    # ---------- Setup / Restart ----------
    def _reset_timer(self):
        if self.timer_job is not None:
            try: self.root.after_cancel(self.timer_job)
            except Exception: pass
            self.timer_job = None
        self.timer_seconds_left = None
        self.game_over = False
        self.timer_label.config(text="00:00")

    def _new_game(self):
        self.moves = 0
        self._update_moves()
        self.rule_break_attempts = 0
        self._update_rule_breaks()
        self.solving = False
        self.interactions_enabled = True
        self._reset_timer()

        # reset move logs
        self.first_click_baseline = None
        self.current_move_start_rel = None
        self.move_logs = []
        self.completed_elapsed_ms = None

        # rebuild board
        self.canvas.delete("all")
        self.pegs = [[], [], []]
        self.disc_sizes.clear()
        self._draw_board()
        # generate pastel colors for current number of discs
        self.palette = self._gen_palette(self.num_discs)
        self._spawn_discs(self.num_discs)
        self._update_min_moves()

        self._refresh_log_window()

    def _change_disks(self, delta):
        new_n = max(self.min_discs, min(self.max_discs, self.num_discs + delta))
        if new_n != self.num_discs:
            self.num_discs = new_n
            self.disks_label_val.config(text=str(self.num_discs))
            self._new_game()

    # ---------- Timer ----------
    def _change_timer_minutes(self, delta):
        new_m = max(0, min(99, self.timer_minutes + delta))
        if new_m != self.timer_minutes:
            self.timer_minutes = new_m
            self.timer_min_label.config(text=str(self.timer_minutes))
            if self.timer_seconds_left is None:
                self.timer_label.config(text="00:00")

    def _start_countdown_if_needed(self):
        if self.timer_minutes >= 1 and self.timer_seconds_left is None:
            self.timer_seconds_left = self.timer_minutes * 60
            self._tick_timer()

    def _tick_timer(self):
        if self.timer_seconds_left is None:
            return
        mins = self.timer_seconds_left // 60
        secs = self.timer_seconds_left % 60
        self.timer_label.config(text=f"{mins:02d}:{secs:02d}")
        if self.timer_seconds_left <= 0:
            self._on_time_up()
            return
        self.timer_seconds_left -= 1
        self.timer_job = self.root.after(1000, self._tick_timer)

    def _on_time_up(self):
        self.game_over = True
        self.interactions_enabled = False
        if self.first_click_baseline is not None:
            now_ms = int((time.perf_counter() - self.first_click_baseline) * 1000)
            self.completed_elapsed_ms = now_ms
            self.move_logs.append((now_ms, now_ms, "-", "GAMEOVER"))
            self._refresh_log_window()

        messagebox.showerror("GAMEOVER", "หมดเวลาแล้ว!")

    # ---------- Drawing ----------
    def _draw_board(self):
        w = self.canvas.winfo_width() or self.canvas.winfo_reqwidth()
        h = self.canvas.winfo_height() or self.canvas.winfo_reqheight()
        margin = 60
        base_y = int(h * 0.8)
        self.base_y = base_y

        self.canvas.create_rectangle(margin//2, base_y, w - margin//2, base_y + 8,
                                     fill=self.base_color, outline=self.base_color, tags=("base",))
        self.peg_x = [int(w*0.2), int(w*0.5), int(w*0.8)]
        self.peg_top = int(h * 0.25)
        self.peg_bottom = base_y

        self.peg_items = []
        for x in self.peg_x:
            peg = self.canvas.create_rectangle(x-5, self.peg_top, x+5, self.peg_bottom,
                                               fill=self.peg_color, outline=self.peg_color, tags=("peg",))
            self.peg_items.append(peg)

    def _spawn_discs(self, n):
        for size in range(n, 0, -1):
            self._create_disc(0, size)

    def _create_disc(self, peg_index, size):
        max_w, min_w = 220, 80
        width = int(min_w + (size-1) * (max_w - min_w) / max(1, self.num_discs - 1))
        height = 22
        x_center = self.peg_x[peg_index]
        y = self._disc_y_for_peg(peg_index)
        left, right = x_center - width//2, x_center + width//2
        color = self.palette[size-1] if size-1 < len(self.palette) else '#cccccc'
        disc = self.canvas.create_rectangle(left, y-height, right, y, fill=color,
                                            outline=color, width=2, tags=("disc",))
        self.pegs[peg_index].append(disc)
        self.disc_sizes[disc] = size
        self._raise_top(peg_index)

    def _disc_y_for_peg(self, peg_index):
        stack_height = len(self.pegs[peg_index])
        return self.base_y - 8 - stack_height * 24

    def _raise_top(self, peg_index):
        for disc in self.pegs[peg_index]:
            self.canvas.tag_raise(disc)

    def _redraw(self):
        self.canvas.delete("all")
        self._draw_board()
        for peg_index, stack in enumerate(self.pegs):
            for level, disc in enumerate(stack[:]):
                size = self.disc_sizes[disc]
                max_w, min_w = 220, 80
                width = int(min_w + (size-1) * (max_w - min_w) / max(1, self.num_discs - 1))
                height = 22
                x_center = self.peg_x[peg_index]
                y = self.base_y - 8 - (len(stack) - level) * 24
                left, right = x_center - width//2, x_center + width//2
                color = self.canvas.itemcget(disc, "fill")
                self.canvas.delete(disc)
                new_disc = self.canvas.create_rectangle(left, y-height, right, y, fill=color,
                                                        outline=color, width=2, tags=("disc",))
                self.disc_sizes[new_disc] = size
                stack[level] = new_disc
        self._update_min_moves()

    # ---------- Drag & Drop ----------
    def _nearest_peg_from_x(self, x):
        distances = [abs(x - xp) for xp in self.peg_x]
        return min(range(3), key=lambda i: distances[i])

    def _on_mouse_down(self, event):
        if not self.interactions_enabled or self.game_over:
            return
        peg = self._nearest_peg_from_x(event.x)
        if not self.pegs[peg]:
            return
        top_disc = self.pegs[peg][-1]
        bbox = self.canvas.bbox(top_disc)
        if bbox and bbox[0] <= event.x <= bbox[2] and bbox[1] <= event.y <= bbox[3]:
            self.dragging_disc = top_disc
            self.drag_from_peg = peg
            self.drag_offset_x = event.x - ((bbox[0]+bbox[2])//2)
            self.drag_offset_y = event.y - ((bbox[1]+bbox[3])//2)
            self.canvas.tag_raise(self.dragging_disc)

            if self.first_click_baseline is None:
                self.first_click_baseline = time.perf_counter()
                self._start_countdown_if_needed()
            if self.first_click_baseline is not None:
                self.current_move_start_rel = time.perf_counter() - self.first_click_baseline

    def _on_mouse_move(self, event):
        if not self.interactions_enabled or self.game_over:
            return
        if self.dragging_disc:
            bbox = self.canvas.bbox(self.dragging_disc)
            cx = (bbox[0]+bbox[2])//2
            cy = (bbox[1]+bbox[3])//2
            dx = event.x - self.drag_offset_x - cx
            dy = event.y - self.drag_offset_y - cy
            self.canvas.move(self.dragging_disc, dx, dy)

    def _on_mouse_up(self, event):
        if not self.interactions_enabled or self.game_over:
            return
        if not self.dragging_disc:
            return

        target_peg = self._nearest_peg_from_x(event.x)
        size = self.disc_sizes[self.dragging_disc]
        valid = False
        attempted_rule_break = False
        if not self.pegs[target_peg]:
            valid = True
        else:
            top_size = self.disc_sizes[self.pegs[target_peg][-1]]
            if size < top_size:
                valid = True
            elif size > top_size:
                attempted_rule_break = True

        if valid:
            if self.pegs[self.drag_from_peg] and self.pegs[self.drag_from_peg][-1] == self.dragging_disc:
                self.pegs[self.drag_from_peg].pop()
            else:
                for i in range(3):
                    if self.dragging_disc in self.pegs[i]:
                        self.pegs[i].remove(self.dragging_disc)
                        break
            self.pegs[target_peg].append(self.dragging_disc)
            self._snap_disc_to_peg(self.dragging_disc, target_peg)
            self.moves += 1
            self._update_moves()
            if self.first_click_baseline is not None and self.current_move_start_rel is not None:
                end_rel = time.perf_counter() - self.first_click_baseline
                self._append_move_log(self.current_move_start_rel, end_rel, self.drag_from_peg, target_peg)
            self._check_win()
        else:
            if attempted_rule_break:
                self.rule_break_attempts += 1
                self._update_rule_breaks()
            self._snap_disc_to_peg(self.dragging_disc, self.drag_from_peg)

        self.dragging_disc = None
        self.drag_from_peg = None
        self.current_move_start_rel = None

    def _snap_disc_to_peg(self, disc, peg_index):
        size = self.disc_sizes[disc]
        max_w, min_w = 220, 80
        width = int(min_w + (size-1) * (max_w - min_w) / max(1, self.num_discs - 1))
        height = 22
        x_center = self.peg_x[peg_index]
        y = self._disc_y_for_peg(peg_index)
        left, right = x_center - width//2, x_center + width//2
        bbox = self.canvas.bbox(disc)
        if not bbox: return
        current_x = (bbox[0]+bbox[2])//2
        current_y = (bbox[1]+bbox[3])//2
        target_x = (left+right)//2
        target_y = y - height//2
        steps = 8
        for i in range(steps):
            dx = (target_x - current_x) / (steps - i)
            dy = (target_y - current_y) / (steps - i)
            self.canvas.move(disc, dx, dy)
            self.canvas.update_idletasks()
            self.canvas.after(8)
            current_x += dx
            current_y += dy
        bbox = self.canvas.bbox(disc)
        dx = target_x - (bbox[0]+bbox[2])//2
        dy = target_y - (bbox[1]+bbox[3])//2
        self.canvas.move(disc, dx, dy)

    # ---------- Game logic ----------
    def _update_moves(self):
        self.moves_label.config(text=str(self.moves))

    def _update_rule_breaks(self):
        self.break_label.config(text=str(self.rule_break_attempts))

    def _update_min_moves(self):
        self.min_moves_label.config(text=f"Minimum Moves: {2 ** self.num_discs - 1}")

    def _check_win(self):
        if len(self.pegs[2]) == self.num_discs and not self.game_over:
            self.interactions_enabled = False
            if self.timer_job is not None:
                try: self.root.after_cancel(self.timer_job)
                except Exception: pass
                self.timer_job = None
            if self.first_click_baseline is not None:
                now_ms = int((time.perf_counter() - self.first_click_baseline) * 1000)
                self.completed_elapsed_ms = now_ms
                self.move_logs.append((now_ms, now_ms, "-", "Success!"))
                self._refresh_log_window()

            messagebox.showinfo("You win!", f"Great job! You solved it in {self.moves} moves.")

    # ---------- Move log ----------
    def _append_move_log(self, start_rel, end_rel, from_peg, to_peg):
        start_ms = int(start_rel * 1000)
        end_ms   = int(end_rel * 1000)
        self.move_logs.append((start_ms, end_ms, from_peg+1, to_peg+1))
        self._refresh_log_window()

    def _open_log_window(self):
        if self.log_window is None or not self._widget_exists(self.log_window):
            self.log_window = tk.Toplevel(self.root)
            self.log_window.title("Move Log")
            self.log_window.configure(bg=self.bg)
            self.log_window.geometry("420x300")
            self.log_window.resizable(True, True)
            self.log_window.protocol("WM_DELETE_WINDOW", self._close_log_window)
            self.log_text = tk.Text(self.log_window, bg="#0b1e2d", fg="#cfe8ff",
                                    font=("Consolas", 12), relief="flat", wrap="none")
            self.log_text.pack(fill="both", expand=True, padx=8, pady=8)
        self._refresh_log_window()
        self.log_window.deiconify(); self.log_window.lift()

    def _close_log_window(self):
        try:
            if self._widget_exists(self.log_window):
                self.log_window.destroy()
        finally:
            self.log_window = None
            self.log_text = None

    def _refresh_log_window(self):
        if self.log_text is None or not self._widget_exists(self.log_text):
            return
        lines = []
        for item in self.move_logs:
            if isinstance(item, tuple) and len(item) == 4:
                a, b, c, d = item
                if c == "-":
                    lines.append(f"{a}, {d}")
                else:
                    lines.append(f"{a}, {b}, {c}, {d}")
            else:
                lines.append(str(item))
        self.log_text.config(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.insert("end", "\n".join(lines))
        self.log_text.config(state="normal")

    @staticmethod
    def _widget_exists(widget):
        try: return bool(widget.winfo_exists())
        except Exception: return False

    # ---------- Save & Table ----------
    def _save_record(self):
        name = self.name_var.get().strip()
        if not name:
            messagebox.showwarning("กรุณาใส่ชื่อ", "โปรดกรอกชื่อผู้เล่นก่อนกด Save")
            self.name_entry.focus_set()
            return
        timer_set_str = f"{self.timer_minutes}:00" if self.timer_minutes >= 1 else "0:00"
        if self.timer_seconds_left is None or self.timer_minutes == 0:
            remaining_str = "0:00"
        else:
            remaining_str = self.timer_label.cget("text")  # "MM:SS"
        elapsed_ms = 0
        if self.first_click_baseline is not None:
            elapsed_ms = self.completed_elapsed_ms if self.completed_elapsed_ms is not None else int((time.perf_counter() - self.first_click_baseline) * 1000)
        record = {
            "Name": name,
            "Num of Disc": self.num_discs,
            "Move": self.moves,
            "Breaking rules": self.rule_break_attempts,
            "Timer": timer_set_str,
            "Remaining time": remaining_str,
            "Time spent(ms)": str(int(elapsed_ms)),
        }
        self.records.append(record)
        messagebox.showinfo("Saved", "บันทึกข้อมูลเรียบร้อย")
        self._refresh_table_window()

    def _open_table_window(self):
        if self.table_window is None or not self._widget_exists(self.table_window):
            self.table_window = tk.Toplevel(self.root)
            self.table_window.title("Statistics")
            self.table_window.configure(bg=self.bg)
            self.table_window.geometry("900x420")
            self.table_window.resizable(True, True)
            self.table_window.protocol("WM_DELETE_WINDOW", self._close_table_window)

            columns = ("Name", "Num of Disc", "Move", "Breaking rules", "Timer", "Remaining time", "Time spent")
            self.table_tree = ttk.Treeview(self.table_window, columns=columns, show="headings")
            for col in columns:
                self.table_tree.heading(col, text=col)
                width = 150 if col == "Name" else 130 if col == "Breaking rules" else 120
                self.table_tree.column(col, anchor="center", width=width)
            self.table_tree.pack(fill="both", expand=True, padx=8, pady=(8,4))

            # Export buttons
            btns = tk.Frame(self.table_window, bg=self.bg)
            btns.pack(fill="x", padx=8, pady=(0,8))
            export_xlsx = self._mk_btn(btns, "Export Excel (.xlsx)", self._export_excel)
            export_csv = self._mk_btn(btns, "Export CSV", self._export_csv)
            export_xlsx.pack(side="left", padx=(0,6))
            export_csv.pack(side="left", padx=(0,6))

        self._refresh_table_window()
        self.table_window.deiconify(); self.table_window.lift()

    def _close_table_window(self):
        try:
            if self._widget_exists(self.table_window):
                self.table_window.destroy()
        finally:
            self.table_window = None
            self.table_tree = None

    def _refresh_table_window(self):
        if self.table_tree is None or not self._widget_exists(self.table_tree):
            return
        for item in self.table_tree.get_children():
            self.table_tree.delete(item)
        for rec in self.records:
            self.table_tree.insert("", "end", values=(
                rec.get("Name", ""),
                rec.get("Num of Disc", ""),
                rec.get("Move", ""),
                rec.get("Breaking rules", 0),
                rec.get("Timer", ""),
                rec.get("Remaining time", ""),
                rec.get("Time spent", "")
            ))

    def _export_csv(self):
        if not self.records:
            messagebox.showinfo("ไม่มีข้อมูล", "ยังไม่มีข้อมูลให้ส่งออก")
            return
        path = filedialog.asksaveasfilename(defaultextension=".csv",
                                            filetypes=[("CSV", "*.csv")],
                                            title="Save CSV")
        if not path:
            return
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Name", "Num of Disc", "Move", "Breaking rules", "Timer", "Remaining time", "Time spent"])
            for r in self.records:
                writer.writerow([
                    r.get("Name", ""),
                    r.get("Num of Disc", ""),
                    r.get("Move", ""),
                    r.get("Breaking rules", 0),
                    r.get("Timer", ""),
                    r.get("Remaining time", ""),
                    r.get("Time spent", "")
                ])
        messagebox.showinfo("Exported", f"ส่งออก CSV สำเร็จ:\n{path}")

    def _export_excel(self):
        if not self.records:
            messagebox.showinfo("ไม่มีข้อมูล", "ยังไม่มีข้อมูลให้ส่งออก")
            return
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font
        except Exception:
            messagebox.showwarning("ต้องการ openpyxl",
                                   "ไม่พบไลบรารี 'openpyxl' จึงจะสร้าง .xlsx ได้\nกำลังเปิดหน้าบันทึก CSV แทน")
            self._export_csv()
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel Workbook", "*.xlsx")],
                                            title="Save Excel")
        if not path:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hanoi Stats"
        headers = ["Name", "Num of Disc", "Move", "Breaking rules", "Timer", "Remaining time", "Time spent"]
        ws.append(headers)
        for r in self.records:
            ws.append([
                r.get("Name", ""),
                r.get("Num of Disc", ""),
                r.get("Move", ""),
                r.get("Breaking rules", 0),
                r.get("Timer", ""),
                r.get("Remaining time", ""),
                r.get("Time spent", "")
            ])
        bold = Font(bold=True)
        for c, _h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c); cell.font = bold
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 22 if headers[c-1] == "Name" else 18
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
        wb.save(path)
        messagebox.showinfo("Exported", f"ส่งออก Excel สำเร็จ:\n{path}")

    # ---------- Solver ----------
    def _solve_animate(self):
        if self.solving: return
        if len(self.pegs[0]) != self.num_discs:
            if not messagebox.askyesno("Restart required", "Solver needs the starting position. Restart now?"):
                return
            self._new_game()
        self.solving = True
        self.interactions_enabled = False
        moves = []
        self._hanoi(self.num_discs, 0, 2, 1, moves)
        self._animate_moves(moves, 0)

    def _hanoi(self, n, src, dst, aux, moves):
        if n == 0: return
        self._hanoi(n-1, src, aux, dst, moves)
        moves.append((src, dst))
        self._hanoi(n-1, aux, dst, src, moves)

    def _animate_moves(self, moves, idx):
        if idx >= len(moves):
            self.solving = False
            self._check_win()
            return
        src, dst = moves[idx]
        if not self.pegs[src]:
            self._animate_moves(moves, idx+1); return
        disc = self.pegs[src][-1]
        self.pegs[src].pop()
        if self.pegs[dst] and self.disc_sizes[self.pegs[dst][-1]] < self.disc_sizes[disc]:
            self.pegs[src].append(disc)
            self._animate_moves(moves, idx+1); return
        self.pegs[dst].append(disc)
        self._snap_disc_to_peg(disc, dst)
        self.moves += 1
        self._update_moves()
        self.canvas.after(120, lambda: self._animate_moves(moves, idx+1))


if __name__ == "__main__":
    root = tk.Tk()
    app = HanoiGUI(root)
    root.mainloop()
